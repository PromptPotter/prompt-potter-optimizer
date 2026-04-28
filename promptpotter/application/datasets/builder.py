"""Ground-truth loaders (Excel, HuggingFace) and train/test splitter.

Each loader returns ``list[Sample]`` — the canonical per-sample domain
object. Legacy dicts are no longer returned; conversion happens at load
time via ``Sample.from_dict`` for any on-disk data predating pass 2.
"""

from __future__ import annotations

import hashlib
import logging
import random
import re
from collections.abc import Callable
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from promptpotter.application.datasets.trace_dataset import load_potter_traces
from promptpotter.domain.sample import Sample
from promptpotter.shared.hashing import HASH_TRUNCATE

if TYPE_CHECKING:
    from promptpotter.domain.pipeline_schema import PipelineSchema
    from promptpotter.domain.search_point import JobSearchPoint

logger = logging.getLogger(__name__)

__all__ = [
    "DATASET_LOADERS",
    "SHEET_COLUMN_MAP",
    "build_dataset_run_data",
    "load_aime_2025",
    "load_bbeh",
    "load_dataset",
    "load_excel_ground_truth",
    "load_gsm8k",
    "load_potter_traces",
    "sample_dataset",
    "samples_from_dicts",
    "split_train_test",
]

# Column mapping per Excel sheet type.
# Keys = sheet names, values = {"query": <col>, "gt": <col>}.
SHEET_COLUMN_MAP: dict[str, dict[str, str]] = {
    "Sheet1": {"query": "Material name in BOM", "gt": "Dataset in SP"},
    "Material": {"query": "Material name in BOM", "gt": "Dataset in SP"},
    "Processes": {"query": "Ausgangsliste", "gt": "Eintrag aus Zielliste"},
    "Processing": {"query": "Ausgangsliste", "gt": "Eintrag aus Zielliste"},
}


def samples_from_dicts(items: list[dict]) -> list[Sample]:
    """Convert a list of dicts (e.g. from on-disk JSON) into Samples.

    Assigns positional ``id`` to items lacking one (legacy migration);
    ``sample_id`` is accepted as ``id``. Extra keys are ignored.
    """
    return [Sample.from_dict(item, fallback_id=i) for i, item in enumerate(items)]


def load_excel_ground_truth(
    path: str | Path,
    sheet_column_map: dict[str, dict[str, str]] = SHEET_COLUMN_MAP,
) -> list[dict]:
    """Read sheets from an Excel file; returns internal staging dicts.

    Staging format carries ``source_sheet`` for downstream partitioning in
    ``split_train_test`` and is not Sample-shaped yet — ids are assigned at
    split time so the global positional index is stable across train/test.
    """
    try:
        import openpyxl
    except ModuleNotFoundError:
        raise ModuleNotFoundError(
            "openpyxl is required to load Excel files. "
            'Install the Excel extras: pip install -e ".[excel]"'
        ) from None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []

    for sheet_name, col_map in sheet_column_map.items():
        if sheet_name not in wb.sheetnames:
            logger.debug("Sheet %r not found in %s, skipping", sheet_name, path)
            continue

        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else "" for h in header_row]

        query_col = col_map["query"]
        gt_col = col_map["gt"]

        if query_col not in headers or gt_col not in headers:
            logger.warning(
                "Sheet %r: expected columns %r and %r but found %s",
                sheet_name,
                query_col,
                gt_col,
                headers,
            )
            continue

        qi = headers.index(query_col)
        gi = headers.index(gt_col)

        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            query = str(row_vals[qi]).strip() if row_vals[qi] is not None else ""
            gt = str(row_vals[gi]).strip() if row_vals[gi] is not None else ""
            if query and gt:
                rows.append(
                    {
                        "query": query,
                        "ground_truth": gt,
                        "source_sheet": sheet_name,
                    }
                )

    wb.close()
    logger.info(
        "Loaded %d ground-truth pairs from %s (%s)",
        len(rows),
        path,
        ", ".join(
            f"{s}: {sum(1 for r in rows if r['source_sheet'] == s)}"
            for s in dict.fromkeys(r["source_sheet"] for r in rows)
        ),
    )
    return rows


def split_train_test(
    data: list[dict],
    test_fraction: float = 0.2,
    seed: int = 42,
) -> tuple[list[Sample], dict[str, list[Sample]]]:
    """Split ground-truth dicts into train + per-domain test Samples.

    Stable ``id`` assignment: ids are assigned sequentially across the
    full input list (post-sort, pre-shuffle) so a sample keeps the same
    id regardless of which split it lands in.
    """
    # Assign ids across the full input list so splits share a stable id space.
    enumerated = [{**row, "id": i} for i, row in enumerate(data)]

    rng = random.Random(seed)
    _proc_sheets = {"Processes", "Processing"}
    processes = [r for r in enumerated if r["source_sheet"] in _proc_sheets]
    material = [r for r in enumerated if r["source_sheet"] not in _proc_sheets]

    def _split(items: list[dict]) -> tuple[list[dict], list[dict]]:
        shuffled = items[:]
        rng.shuffle(shuffled)
        n_test = max(1, round(len(shuffled) * test_fraction))
        return shuffled[n_test:], shuffled[:n_test]

    train_proc, test_proc = _split(processes) if processes else ([], [])
    train_mat, test_mat = _split(material) if material else ([], [])

    train_dicts = train_proc + train_mat

    test_queries_proc = {r["query"] for r in test_proc}
    test_queries_mat = {r["query"] for r in test_mat}
    overlap = test_queries_proc & test_queries_mat
    if overlap:
        logger.warning(
            "Duplicate queries across test sets (%d): %s",
            len(overlap),
            list(overlap)[:5],
        )

    train = [Sample.from_dict(d) for d in train_dicts]
    test_sets = {
        "test_processes": [Sample.from_dict(d) for d in test_proc],
        "test_material": [Sample.from_dict(d) for d in test_mat],
    }

    logger.info(
        "Split: %d train, %d test_processes, %d test_material",
        len(train),
        len(test_sets["test_processes"]),
        len(test_sets["test_material"]),
    )
    return train, test_sets


def sample_dataset(dataset: list[Sample], sample_size: int) -> list[Sample]:
    """Top-``sample_size`` slice of eval samples.

    Datasets are already shuffled at creation (train/test split, HF load),
    so the loop-time slice is the deterministic prefix — no second RNG.
    """
    if sample_size <= 0:
        raise ValueError(f"sp_budget_ttest must be > 0, got {sample_size}")
    return dataset[:sample_size]


_GSM8K_ANSWER_RE = re.compile(r"####\s*(-?[\d,]+\.?\d*)")


def load_gsm8k(
    split: str = "train",
    sample_size: int = 0,
    seed: int = 42,
) -> list[Sample]:
    """Load GSM8K from HuggingFace.

    Requires the ``datasets`` library: ``pip install -e ".[benchmarks]"``.
    """
    try:
        from datasets import load_dataset  # type: ignore[import-not-found,import-untyped]
    except ModuleNotFoundError:
        raise ModuleNotFoundError(
            "The 'datasets' library is required for GSM8K. "
            'Install the benchmarks extras: pip install -e ".[benchmarks]"'
        ) from None

    ds = load_dataset("openai/gsm8k", "main", split=split)
    samples: list[Sample] = []
    for i, row in enumerate(ds):
        m = _GSM8K_ANSWER_RE.search(row["answer"])
        gt = f"#### {m.group(1)}" if m else row["answer"].strip()
        samples.append(Sample(id=i, query=row["question"], ground_truth=gt))

    if sample_size > 0 and len(samples) > sample_size:
        samples = random.Random(seed).sample(samples, sample_size)

    logger.info("Loaded GSM8K %s: %d items", split, len(samples))
    return samples


def load_aime_2025(
    sample_size: int = 0,
    seed: int = 42,
) -> list[Sample]:
    """Load AIME 2025 from HuggingFace.

    Requires the ``datasets`` library: ``pip install -e ".[benchmarks]"``.
    """
    try:
        from datasets import load_dataset  # type: ignore[import-not-found,import-untyped]
    except ModuleNotFoundError:
        raise ModuleNotFoundError(
            "The 'datasets' library is required for AIME 2025. "
            'Install the benchmarks extras: pip install -e ".[benchmarks]"'
        ) from None

    ds = load_dataset("MathArena/aime_2025", split="train")
    samples: list[Sample] = [
        Sample(id=i, query=row["problem"], ground_truth=str(row["answer"]))
        for i, row in enumerate(ds)
    ]

    if sample_size > 0 and len(samples) > sample_size:
        samples = random.Random(seed).sample(samples, sample_size)

    logger.info("Loaded AIME 2025: %d items", len(samples))
    return samples


# ---------------------------------------------------------------------------
# Dataset loader registry — add new loaders here
# ---------------------------------------------------------------------------


def load_bbeh(sample_size: int = 0, seed: int = 42) -> list[Sample]:
    """Load BBEH mini (460 examples, 23 tasks) from HuggingFace.

    BBEH ships as 23 task bins with no native per-sample ID; we assign
    sequential ``Sample.id`` after flattening. The per-task metadata is
    dropped (no consumers after pass 2).
    """
    try:
        from datasets import load_dataset  # type: ignore[import-not-found,import-untyped]
    except ModuleNotFoundError:
        raise ModuleNotFoundError(
            "The 'datasets' library is required for BBEH. "
            'Install the benchmarks extras: pip install -e ".[benchmarks]"'
        ) from None

    ds = load_dataset("BBEH/bbeh", split="train")
    samples: list[Sample] = []
    for row in ds:
        if not row.get("mini"):
            continue
        samples.append(
            Sample(
                id=len(samples),
                query=row["input"],
                ground_truth=row["target"],
            )
        )

    if sample_size > 0 and len(samples) > sample_size:
        samples = random.Random(seed).sample(samples, sample_size)

    logger.info("Loaded BBEH mini: %d items", len(samples))
    return samples


DATASET_LOADERS: dict[str, Callable[..., list[Sample]]] = {
    "gsm8k": load_gsm8k,
    "aime_2025": load_aime_2025,
    "bbeh": load_bbeh,
}
"""Map dataset name → loader function. Register new datasets here.

``load_potter_traces`` is a specialized meta-optimization replay loader
whose rows carry extra fields (round_context, score_delta, …) beyond
Sample's schema. It is available via direct import but not routed
through the Sample-typed registry.
"""


def load_dataset(name: str, **kwargs: Any) -> list[Sample]:
    """Load a dataset by name from the registry.

    Raises :class:`KeyError` for unknown dataset names.
    """
    loader = DATASET_LOADERS.get(name)
    if loader is None:
        raise KeyError(f"Unknown dataset {name!r}. Available: {sorted(DATASET_LOADERS)}")
    return loader(**kwargs)


def build_dataset_run_data(
    run_id: str,
    name: str,
    content_hash: str,
    search_point: JobSearchPoint,
    scores: dict,
    results: list,
    *,
    source: str = "",
    experiment_id: str = "",
    pipeline_schema: PipelineSchema | None = None,
) -> dict:
    """Build a measurement-batch dict ready for ``Stores.archive.save()``."""
    rendered_prompt = search_point.render()
    sp_h = search_point.sp_hash(pipeline_schema)
    data: dict = {
        "run_id": run_id,
        "name": name,
        "content_hash": content_hash,
        "prompt_fields_id": sp_h,
        "rendered_prompt_hash": hashlib.sha256(
            rendered_prompt.encode(),
        ).hexdigest()[:HASH_TRUNCATE],
        "item_count": scores["total"],
        "scores": scores,
        "source": source,
        "created_at": datetime.now(UTC).isoformat(),
        "measurements": list(results),
    }
    if pipeline_schema and search_point.pipeline_params:
        data["node_configs"] = pipeline_schema.node_configs(search_point.pipeline_params)
    if search_point.pipeline_params:
        data["pipeline_params"] = search_point.pipeline_params
    if experiment_id:
        data["experiment_id"] = experiment_id
    return data
