"""Ground-truth loaders (Excel, HuggingFace, traces) and train/test splitter.

Each loader returns ``[{"query": str, "ground_truth": str, ...}]``.
"""

from __future__ import annotations

import hashlib
import logging
import random
import re
from collections.abc import Callable
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from promptpotter.shared.hashing import HASH_TRUNCATE

if TYPE_CHECKING:
    from promptpotter.models.pipeline_schema import PipelineSchema
    from promptpotter.models.search_point import JobSearchPoint
    from promptpotter.services.project_store import ProjectStore

logger = logging.getLogger(__name__)

__all__ = [
    "DATASET_LOADERS",
    "SHEET_COLUMN_MAP",
    "build_dataset_run_data",
    "load_aime_2025",
    "load_dataset",
    "load_dataset_from_traces",
    "load_excel_ground_truth",
    "load_gsm8k",
    "sample_dataset",
    "split_train_test",
]

# Column mapping per Excel sheet type.
# Keys = sheet names, values = {"query": <col>, "gt": <col>}.
SHEET_COLUMN_MAP: dict[str, dict[str, str]] = {
    "Sheet1": {"query": "Material name in BOM", "gt": "Dataset in SP"},
    "Material": {"query": "Material name in BOM", "gt": "Dataset in SP"},
    "Processes": {"query": "Ausgangsliste", "gt": "Eintrag aus Zielliste"},
    "Processing": {"query": "Ausgangsliste", "gt": "Eintrag aus Zielliste"},
}


def load_excel_ground_truth(
    path: str | Path,
    sheet_column_map: dict[str, dict[str, str]] = SHEET_COLUMN_MAP,
) -> list[dict]:
    """Read sheets from an Excel file, normalize columns, return unified list.

    Returns:
        List of ``{"query": str, "ground_truth": str, "source_sheet": str}``.
        Rows where query or ground_truth is empty/NaN are dropped.
    """
    try:
        import openpyxl
    except ModuleNotFoundError:
        raise ModuleNotFoundError(
            "openpyxl is required to load Excel files. Install it: pip install openpyxl"
        ) from None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []

    for sheet_name, col_map in sheet_column_map.items():
        if sheet_name not in wb.sheetnames:
            logger.debug("Sheet %r not found in %s, skipping", sheet_name, path)
            continue

        ws = wb[sheet_name]
        # Build header index from first row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h else "" for h in header_row]

        query_col = col_map["query"]
        gt_col = col_map["gt"]

        if query_col not in headers or gt_col not in headers:
            logger.warning(
                "Sheet %r: expected columns %r and %r but found %s",
                sheet_name,
                query_col,
                gt_col,
                headers,
            )
            continue

        qi = headers.index(query_col)
        gi = headers.index(gt_col)

        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            query = str(row_vals[qi]).strip() if row_vals[qi] is not None else ""
            gt = str(row_vals[gi]).strip() if row_vals[gi] is not None else ""
            if query and gt:
                rows.append(
                    {
                        "query": query,
                        "ground_truth": gt,
                        "source_sheet": sheet_name,
                    }
                )

    wb.close()
    logger.info(
        "Loaded %d ground-truth pairs from %s (%s)",
        len(rows),
        path,
        ", ".join(
            f"{s}: {sum(1 for r in rows if r['source_sheet'] == s)}"
            for s in dict.fromkeys(r["source_sheet"] for r in rows)
        ),
    )
    return rows


def split_train_test(
    data: list[dict],
    test_fraction: float = 0.2,
    seed: int = 42,
) -> tuple[list[dict], dict[str, list[dict]]]:
    """Split ground-truth data into train + per-domain test sets.

    Logic:
    - Separate rows by source_sheet: "Processes" vs ("Material" + "Sheet1")
    - From each group, draw ``test_fraction`` into the test set
    - Remaining rows from both groups -> combined train set

    Returns:
        ``(train_data, {"test_processes": [...], "test_material": [...]})``.
    """
    rng = random.Random(seed)

    _proc_sheets = {"Processes", "Processing"}
    processes = [r for r in data if r["source_sheet"] in _proc_sheets]
    material = [r for r in data if r["source_sheet"] not in _proc_sheets]

    def _split(items: list[dict]) -> tuple[list[dict], list[dict]]:
        shuffled = items[:]
        rng.shuffle(shuffled)
        n_test = max(1, round(len(shuffled) * test_fraction))
        return shuffled[n_test:], shuffled[:n_test]

    train_proc, test_proc = _split(processes) if processes else ([], [])
    train_mat, test_mat = _split(material) if material else ([], [])

    train = train_proc + train_mat

    # Warn on duplicate queries across test sets
    test_queries_proc = {r["query"] for r in test_proc}
    test_queries_mat = {r["query"] for r in test_mat}
    overlap = test_queries_proc & test_queries_mat
    if overlap:
        logger.warning(
            "Duplicate queries across test sets (%d): %s",
            len(overlap),
            list(overlap)[:5],
        )

    logger.info(
        "Split: %d train, %d test_processes, %d test_material",
        len(train),
        len(test_proc),
        len(test_mat),
    )
    return train, {"test_processes": test_proc, "test_material": test_mat}


def sample_dataset(
    dataset: list[dict],
    sample_size: int,
    seed: int = 42,
) -> list[dict]:
    """Deterministic subsample of eval queries.

    ``sample_size`` must be a positive integer.  Returns the full list
    unchanged only when the dataset is already at or below that size.
    """
    if sample_size <= 0:
        raise ValueError(f"sp_budget_ttest must be > 0, got {sample_size}")
    if len(dataset) <= sample_size:
        return dataset
    return random.Random(seed).sample(dataset, sample_size)


_GSM8K_ANSWER_RE = re.compile(r"####\s*(-?[\d,]+\.?\d*)")


def load_gsm8k(
    split: str = "train",
    sample_size: int = 0,
    seed: int = 42,
) -> list[dict]:
    """Load GSM8K from HuggingFace and return DatasetStore-format items.

    Returns:
        List of ``{"query": str, "ground_truth": str}`` dicts.
        ``ground_truth`` is the ``#### N`` answer line only.

    Requires the ``datasets`` library: ``pip install datasets``.
    """
    try:
        from datasets import load_dataset  # type: ignore[import-untyped]
    except ModuleNotFoundError:
        raise ModuleNotFoundError(
            "The 'datasets' library is required for GSM8K. Install it: pip install datasets"
        ) from None

    ds = load_dataset("openai/gsm8k", "main", split=split)
    items: list[dict] = []
    for row in ds:
        # Extract just the "#### N" answer line from the full solution
        m = _GSM8K_ANSWER_RE.search(row["answer"])
        gt = f"#### {m.group(1)}" if m else row["answer"].strip()
        items.append({"query": row["question"], "ground_truth": gt})

    if sample_size > 0 and len(items) > sample_size:
        items = random.Random(seed).sample(items, sample_size)

    logger.info("Loaded GSM8K %s: %d items", split, len(items))
    return items


def load_aime_2025(
    sample_size: int = 0,
    seed: int = 42,
) -> list[dict]:
    """Load AIME 2025 from HuggingFace and return DatasetStore-format items.

    Returns:
        List of ``{"query": str, "ground_truth": str}`` dicts.
        ``ground_truth`` is the integer answer as a string.

    Requires the ``datasets`` library: ``pip install datasets``.
    """
    try:
        from datasets import load_dataset  # type: ignore[import-untyped]
    except ModuleNotFoundError:
        raise ModuleNotFoundError(
            "The 'datasets' library is required for AIME 2025. Install it: pip install datasets"
        ) from None

    ds = load_dataset("MathArena/aime_2025", split="train")
    items: list[dict] = []
    for row in ds:
        items.append({"query": row["problem"], "ground_truth": str(row["answer"])})

    if sample_size > 0 and len(items) > sample_size:
        items = random.Random(seed).sample(items, sample_size)

    logger.info("Loaded AIME 2025: %d items", len(items))
    return items


# ---------------------------------------------------------------------------
# Dataset loader registry — add new loaders here
# ---------------------------------------------------------------------------

DATASET_LOADERS: dict[str, Callable[..., list[dict]]] = {
    "gsm8k": load_gsm8k,
    "aime_2025": load_aime_2025,
}
"""Map dataset name → loader function. Register new datasets here."""


def load_dataset(name: str, **kwargs: Any) -> list[dict]:
    """Load a dataset by name from the registry.

    Raises :class:`KeyError` for unknown dataset names.
    """
    loader = DATASET_LOADERS.get(name)
    if loader is None:
        raise KeyError(f"Unknown dataset {name!r}. Available: {sorted(DATASET_LOADERS)}")
    return loader(**kwargs)


def build_dataset_run_data(
    run_id: str,
    name: str,
    content_hash: str,
    search_point: JobSearchPoint,
    scores: dict,
    results: list,
    *,
    source: str = "",
    experiment_id: str = "",
    prompt_node_names: list[str] | None = None,
) -> dict:
    """Build a DatasetRun dict ready for ProjectStore.save_dataset_run()."""
    rendered_prompt = search_point.render()
    sp_h = search_point.sp_hash(prompt_node_names)
    data: dict = {
        "run_id": run_id,
        "name": name,
        "content_hash": content_hash,
        "prompt_fields_id": sp_h,
        "rendered_prompt_hash": hashlib.sha256(
            rendered_prompt.encode(),
        ).hexdigest()[:HASH_TRUNCATE],
        "item_count": scores["total"],
        "scores": scores,
        "source": source,
        "created_at": datetime.now(UTC).isoformat(),
        "dataset_run_items": [
            {k: v for k, v in r.items() if k != "precomputed_through"} for r in results
        ],
    }
    data["sp_hash"] = sp_h
    if search_point.pipeline_params:
        data["pipeline_params"] = search_point.pipeline_params
    if experiment_id:
        data["experiment_id"] = experiment_id
    return data


# ---------------------------------------------------------------------------
# Trace-based dataset extraction (from synced backend experiments)
# ---------------------------------------------------------------------------


def _generic_resolve_gt(exp_data: dict) -> dict[str, str]:
    """Build ``{query: ground_truth}`` from evaluation_results (generic fallback)."""
    gt_map: dict[str, str] = {}
    runs = exp_data.get("runs", [])
    if not runs:
        return gt_map
    for er in runs[0].get("evaluation_results", []):
        q = er.get("query", "")
        gt = er.get("ground_truth", "")
        if q and gt:
            gt_map[q] = gt
    return gt_map


def _extract_dataset_from_traces(
    exp_data: dict,
    schema: PipelineSchema,
    backend_name: str | None = None,
) -> list:
    """Build dataset from Langfuse-style traces in synced experiment data.

    Each trace in ``runs[0].traces[]`` carries named observations with
    pipeline step outputs.  Ground truth is resolved via a registered
    connector resolver (or generic query-string matching as fallback).

    Observation extraction is driven by the schema's ``obs_extraction_map()``.

    Returns:
        List of dataset dicts (may be empty).
    """
    from promptpotter.config.extractors import TRACE_GT_RESOLVERS

    resolver = TRACE_GT_RESOLVERS.get((backend_name or "").lower())

    # Pre-build generic fallback map if no connector resolver
    generic_gt = _generic_resolve_gt(exp_data) if not resolver else {}

    runs = exp_data.get("runs", [])
    if not runs:
        return []

    traces = runs[0].get("traces", [])
    if not traces:
        return []

    dataset = []
    for trace in traces:
        query = (trace.get("input") or {}).get("query", "")
        if not query:
            continue

        ground_truth = resolver(exp_data, query) if resolver else generic_gt.get(query)
        if not ground_truth:
            continue

        # Build pipeline_data from observations using declarative mapping
        pipeline_data: dict = {}
        llm_provider: str | None = None

        _obs_map = schema.obs_extraction_map()

        for obs in trace.get("observations", []):
            name = obs.get("name", "")
            output = obs.get("output")
            if not name or not output:
                continue

            for field in _obs_map.get(name, []):
                if field.output_field is None:
                    pipeline_data[field.pipeline_key] = output
                else:
                    val = output.get(field.output_field)
                    if val is not None:
                        pipeline_data[field.pipeline_key] = val

                if field.is_llm and not llm_provider:
                    llm_provider = (obs.get("metadata") or {}).get("model")

        # Gate: required pipeline key must be present
        if not pipeline_data.get(schema.required_pipeline_key()):
            continue

        if llm_provider:
            pipeline_data["llm_provider"] = llm_provider

        # Total time from trace scores (latency_ms → seconds)
        for score in trace.get("scores", []):
            if score.get("name") == "latency_ms" and score.get("value"):
                pipeline_data["total_time"] = score["value"] / 1000.0
                break

        dataset.append(
            {
                "query": query,
                "ground_truth": ground_truth,
                "pipeline_data": pipeline_data,
                "status": "success",
            }
        )

    return dataset


def load_dataset_from_traces(
    store: ProjectStore,
    backend_id: str,
    experiment_id: str,
    sample_size: int = 0,
    schema: PipelineSchema | None = None,
) -> list:
    """Load per-query data from synced experiments or stored replays.

    Priority:
        1. Langfuse-style traces from ``runs[0].traces[]``
        2. Stored replay executions
        3. Empty list if neither found
    """
    exp_data = store.backends.load_sync(
        backend_id,
        f"experiments/{experiment_id}.json",
    )

    if exp_data:
        if schema is None:
            from promptpotter.models.pipeline_schema import PipelineSchema

            schema = PipelineSchema()
        backend = store.backends.get(backend_id)
        backend_name = backend.name if backend else None
        dataset = _extract_dataset_from_traces(exp_data, schema=schema, backend_name=backend_name)
        if dataset:
            if sample_size > 0 and len(dataset) > sample_size:
                rng = random.Random(42)
                dataset = rng.sample(dataset, sample_size)
            return dataset

    executions = store.backends.list_executions(backend_id)
    for ex_summary in executions:
        if ex_summary.get("experiment_id") == experiment_id:
            execution = store.backends.load_execution(
                backend_id,
                ex_summary["execution_id"],
            )
            if execution and schema is not None:
                req_key = schema.required_pipeline_key()
                dataset = [
                    r.model_dump()
                    for r in execution.results
                    if r.status == "success" and r.pipeline_data and r.pipeline_data.get(req_key)
                ]
                if dataset:
                    if sample_size > 0 and len(dataset) > sample_size:
                        rng = random.Random(42)
                        dataset = rng.sample(dataset, sample_size)
                    return dataset

    return []
